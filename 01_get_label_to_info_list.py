# KF 12/14/2018

# Read information from the given xlsx file, extract stuctured information and save as an .npy file.

import numpy as np
from openpyxl import load_workbook

xl_file = '20181214_更新_病虫害数据（61类）标签说明.xlsx'

wb = load_workbook(xl_file)
ws = wb.active


# Main loop
main_list = []
for i in range(len(ws['A'])-1):
    label = ws.cell(row=i+2, column=1).value.replace('\xa0', '')
    label = int(label)
    plant = ws.cell(row=i+2, column=6).value
    disease = ws.cell(row=i+2,column=7).value
    severity = ws.cell(row=i+2, column=8).value
    tmp = (label, plant, disease, severity)
    main_list.append(tmp)

wb.close()
np.save('label_plant_disease_severity_zh.npy', main_list)

# Create txt for database table "model_label_dictionary"
# Database fields:
# -------------------------------------------------------------------------------
# label_id | article_id | plant_name | health_status | disease_name | severity
# -------------------------------------------------------------------------------
#      int |       None |        str |      int(0/1) |     str/None | int(0 or 1)/None
# -------------------------------------------------------------------------------
# Note 
database_list = []
for line in main_list:
    label_id = line[0]
    article_id = '' 
    plant_name = line[1]
    health_status = 0
    disease_name = ''
    severity = ''
    if line[2] != '健康':
        health_status = 1
        disease_name = line[2]
        severity = line[3]
    tmp = (label_id, article_id, plant_name, health_status, disease_name, severity)
    database_list.append(tmp)
    #print(tmp)

np.savetxt('data_for_model_label_dictionary.txt', database_list, fmt='%s', delimiter=',')

# Create txt for database table "plant_dictionary"
plant_set = set()
for line in main_list:
    plant_set.add(line[1])

np.savetxt('data_for_plant_dictionary.txt', list(plant_set), fmt='%s', delimiter=',')

# Create txt for database table "disease_dictionary"
disease_set = set()
for line in main_list:
    if line[2] != '健康':
        disease_set.add(line[2])

np.savetxt('data_for_disease_dictionary.txt', list(disease_set), fmt='%s', delimiter=',')
